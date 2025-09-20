import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re

# Читаем исходный файл products
products_df = pd.read_excel('products.xlsx', header=None)

# Загружаем template для редактирования
template_wb = load_workbook('template.xlsx')
sheet1 = template_wb['Sheet1']

# Функция для более точного извлечения параметров
def extract_all_parameters(description):
    params = {
        'length': None,
        'width': None,
        'height': None,
        'weight': None,
        'package_length': None,
        'package_width': None,
        'package_height': None,
        'package_weight': None,
        'color': None,
        'material': None,
        'type': None,
        'connection': None
    }

    if pd.isna(description):
        return params

    desc = str(description)

    # Извлекаем длину трубки/лейки
    length_patterns = [
        r'Длина трубки:\s*(\d+)\s*мм',
        r'Длина:\s*(\d+)\s*мм',
        r'макс\.\s*(\d+)\s*мм'
    ]
    for pattern in length_patterns:
        match = re.search(pattern, desc)
        if match:
            params['length'] = int(match.group(1))
            break

    # Извлекаем ширину лейки
    width_patterns = [
        r'Ширина лейки:\s*(\d+)\s*мм',
        r'Ширина:\s*(\d+)\s*мм'
    ]
    for pattern in width_patterns:
        match = re.search(pattern, desc)
        if match:
            params['width'] = int(match.group(1))
            break

    # Извлекаем высоту держателя
    height_patterns = [
        r'Высота держателя:\s*(\d+)\s*мм',
        r'Высота:\s*(\d+)\s*мм'
    ]
    for pattern in height_patterns:
        match = re.search(pattern, desc)
        if match:
            params['height'] = int(match.group(1))
            break

    # Извлекаем размеры упаковки
    package_match = re.search(r'Размер упаковки[:\s]*(\d+)[х×](\d+)[х×](\d+)', desc)
    if package_match:
        params['package_length'] = int(package_match.group(1))
        params['package_width'] = int(package_match.group(2))
        params['package_height'] = int(package_match.group(3))

    # Извлекаем вес товара
    weight_match = re.search(r'Вес товара:\s*(\d+)\s*г', desc)
    if weight_match:
        params['package_weight'] = int(weight_match.group(1))

    # Определяем тип полива
    if 'веерный' in desc.lower():
        params['type'] = 'Веерный'
    elif 'дождевой' in desc.lower():
        params['type'] = 'Дождевой'
    elif 'регулируемый' in desc.lower():
        params['type'] = 'Регулируемый'
    elif 'капельный' in desc.lower():
        params['type'] = 'Капельный'
    else:
        params['type'] = 'Универсальный'

    # Определяем цвет
    colors_map = {
        'зеленый': 'Зеленый',
        'зелёный': 'Зеленый',
        'синий': 'Синий',
        'белый': 'Белый',
        'серый': 'Серый',
        'оранжевый': 'Оранжевый'
    }

    desc_lower = desc.lower()
    for color_key, color_value in colors_map.items():
        if color_key in desc_lower:
            params['color'] = color_value
            break

    # Определяем материал
    if 'пластик' in desc_lower:
        params['material'] = 'Пластик'
    elif 'металл' in desc_lower:
        params['material'] = 'Металл'
    elif 'латунь' in desc_lower:
        params['material'] = 'Латунь'
    elif 'алюминий' in desc_lower:
        params['material'] = 'Алюминий'
    else:
        params['material'] = 'Пластик'

    # Определяем соединение
    connection_match = re.search(r'Присоединительный размер:\s*([^\\n]+)', desc)
    if connection_match:
        params['connection'] = connection_match.group(1).strip()

    return params

# Функция для генерации описания товара
def generate_product_description(product_name, model_name, params):
    """Генерирует краткое и информативное описание товара"""

    if pd.isna(product_name):
        return None

    name_lower = str(product_name).lower()

    # Базовые описания для разных типов товаров
    if 'распылитель' in name_lower and 'веерный' in name_lower:
        desc = f"Веерный распылитель для равномерного полива газонов и клумб. "
        desc += f"Создает мягкий веерообразный поток воды, идеальный для деликатного полива растений. "
        if params['material']:
            desc += f"Изготовлен из прочного {params['material'].lower()}а. "
        desc += "Простое подключение к садовому шлангу."

    elif 'распылитель' in name_lower and 'импульсный' in name_lower:
        desc = f"Импульсный распылитель с регулируемой дальностью полива. "
        desc += "Обеспечивает круговое орошение больших площадей. "
        if params['material']:
            desc += f"Корпус из надежного {params['material'].lower()}а. "
        desc += "Идеален для полива газонов и садовых участков."

    elif 'распылитель' in name_lower and model_name and 'дождь' in str(model_name).lower():
        desc = f"Дождевой распылитель для мягкого полива растений. "
        desc += "Имитирует естественный дождь, не повреждая нежные побеги. "
        if params['length']:
            desc += f"Радиус полива до {params['length']} мм. "
        desc += "Подходит для теплиц и цветников."

    elif 'насадка' in name_lower:
        desc = f"Многофункциональная насадка для полива с регулировкой напора. "
        desc += "Различные режимы распыления от тонкой струи до мягкого душа. "
        if params['material']:
            desc += f"Эргономичная рукоятка из {params['material'].lower()}а. "
        desc += "Универсальное соединение 1/2\"."

    elif 'дождеватель' in name_lower and 'круговой' in name_lower:
        desc = f"Круговой дождеватель для автоматического полива. "
        desc += "Вращающаяся конструкция обеспечивает равномерное орошение по кругу. "
        if params['width']:
            desc += f"Диаметр орошения до {params['width']*2} мм. "
        desc += "Регулируемый угол распыления."

    elif 'дождеватель' in name_lower and 'осциллирующий' in name_lower:
        desc = f"Осциллирующий дождеватель для прямоугольных участков. "
        desc += "Маятниковый механизм для равномерного покрытия площади. "
        if params['length'] and params['width']:
            desc += f"Площадь полива до {params['length']}x{params['width']} мм. "
        desc += "Регулируемая ширина полива."

    elif 'пистолет' in name_lower:
        desc = f"Поливочный пистолет с плавной регулировкой потока. "
        desc += "Удобная прорезиненная рукоятка, фиксатор курка для длительного полива. "
        if params['type']:
            desc += f"{params['type']} тип распыления. "
        desc += "Подходит для всех видов садовых работ."

    else:
        # Универсальное описание
        desc = f"Садовое поливочное устройство для эффективного орошения. "
        if params['type']:
            desc += f"{params['type']} тип полива. "
        if params['material']:
            desc += f"Изготовлено из качественного {params['material'].lower()}а. "
        desc += "Надежное и долговечное решение для вашего сада."

    return desc

# Начинаем заполнение со строки 3
row_num = 3

# Обрабатываем данные из products
for idx in range(3, len(products_df)):
    product_row = products_df.iloc[idx]

    # Пропускаем пустые строки
    if pd.isna(product_row[2]):
        continue

    # Извлекаем все параметры из описания
    params = extract_all_parameters(product_row[4])

    # A - Артикул
    if pd.notna(product_row[5]):
        sheet1.cell(row=row_num, column=1, value=str(product_row[5]))

    # B - Наименование товара
    if pd.notna(product_row[2]):
        sheet1.cell(row=row_num, column=2, value=str(product_row[2]))

    # C - Название модели
    if pd.notna(product_row[3]):
        sheet1.cell(row=row_num, column=3, value=str(product_row[3]))

    # D - Краткое описание товара (генерируем)
    description = generate_product_description(product_row[2], product_row[3], params)
    if description:
        sheet1.cell(row=row_num, column=4, value=description)

    # E - Страна производства
    sheet1.cell(row=row_num, column=5, value="Россия")

    # F - Тип распылителя
    if params['type']:
        sheet1.cell(row=row_num, column=6, value=params['type'])

    # G - Материал
    if params['material']:
        sheet1.cell(row=row_num, column=7, value=params['material'])

    # H - Размер соединения
    if params['connection']:
        sheet1.cell(row=row_num, column=8, value=params['connection'])
    else:
        sheet1.cell(row=row_num, column=8, value="1/2\"")

    # I - Единица измерения
    sheet1.cell(row=row_num, column=9, value="шт")

    # J - Длина, мм
    if params['length']:
        sheet1.cell(row=row_num, column=10, value=params['length'])

    # K - Ширина, мм
    if params['width']:
        sheet1.cell(row=row_num, column=11, value=params['width'])

    # L - Высота, мм
    if params['height']:
        sheet1.cell(row=row_num, column=12, value=params['height'])

    # M - Вес, гр
    if params['package_weight']:
        sheet1.cell(row=row_num, column=13, value=params['package_weight'])

    # N - Длина упаковки, мм
    if params['package_length']:
        sheet1.cell(row=row_num, column=14, value=params['package_length'])

    # O - Ширина упаковки, мм
    if params['package_width']:
        sheet1.cell(row=row_num, column=15, value=params['package_width'])

    # P - Высота упаковки, мм
    if params['package_height']:
        sheet1.cell(row=row_num, column=16, value=params['package_height'])

    # Q - Вес упаковки, гр
    if params['package_weight']:
        sheet1.cell(row=row_num, column=17, value=params['package_weight'])

    # R - Поливать дома или на улице
    sheet1.cell(row=row_num, column=18, value="да")

    # S - Тип изделия
    product_name_lower = str(product_row[2]).lower() if pd.notna(product_row[2]) else ""
    if 'распылитель' in product_name_lower:
        sheet1.cell(row=row_num, column=19, value="Распылитель")
    elif 'дождеватель' in product_name_lower:
        sheet1.cell(row=row_num, column=19, value="Дождеватель")
    elif 'насадка' in product_name_lower:
        sheet1.cell(row=row_num, column=19, value="Насадка")
    elif 'пистолет' in product_name_lower:
        sheet1.cell(row=row_num, column=19, value="Пистолет")
    else:
        sheet1.cell(row=row_num, column=19, value="Поливочное устройство")

    # T - Цвет
    if params['color']:
        sheet1.cell(row=row_num, column=20, value=params['color'])

    # U - Штрих-код (оставляем пустым или можно генерировать)
    # sheet1.cell(row=row_num, column=21, value="")

    row_num += 1

# Сохраняем результат
template_wb.save('template_filled_correct.xlsx')

print(f"Данные успешно перенесены и отформатированы!")
print(f"Заполнено {row_num - 3} строк товаров")
print(f"Результат сохранен в файл: template_filled_correct.xlsx")
print("\nЧто было сделано:")
print("- Извлечены числовые параметры из описаний (длина, ширина, высота, вес)")
print("- Сгенерированы информативные описания товаров")
print("- Правильно распределены данные по колонкам")
print("- Определены типы изделий, материалы и цвета")