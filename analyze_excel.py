import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Анализируем файл products.xlsx
print("=== Анализ products.xlsx ===")
products_df = pd.read_excel('products.xlsx', sheet_name=None)
for sheet_name, df in products_df.items():
    print(f"\nЛист '{sheet_name}':")
    print(f"Размер: {df.shape}")
    print(f"Колонки: {list(df.columns)}")
    print("\nПервые 5 строк:")
    print(df.head())

# Анализируем файл template.xlsx
print("\n\n=== Анализ template.xlsx ===")
template_wb = load_workbook('template.xlsx')
print(f"Листы: {template_wb.sheetnames}")

for sheet_name in template_wb.sheetnames:
    sheet = template_wb[sheet_name]
    print(f"\nЛист '{sheet_name}':")
    print(f"Максимальная строка: {sheet.max_row}")
    print(f"Максимальная колонка: {sheet.max_column}")

    # Показываем первые 10 строк
    print("\nСодержимое первых ячеек:")
    for row in range(1, min(11, sheet.max_row + 1)):
        row_data = []
        for col in range(1, min(6, sheet.max_column + 1)):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                row_data.append(f"{cell.coordinate}: {cell.value}")
        if row_data:
            print(f"Строка {row}: {', '.join(row_data)}")