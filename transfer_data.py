import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re

# Читаем исходный файл products
products_df = pd.read_excel('products.xlsx', header=None)

# Загружаем template для редактирования
template_wb = load_workbook('template.xlsx')
sheet1 = template_wb['Sheet1']

# Функция для извлечения размеров из описания
def extract_dimensions(description):
    dimensions = {
        'length': None,
        'width': None,
        'height': None,
        'weight': None,
        'package_length': None,
        'package_width': None,
        'package_height': None,
        'package_weight': None
    }

    if pd.isna(description):
        return dimensions

    description = str(description)

    # Извлекаем длину трубки
    length_match = re.search(r'Длина трубки:\s*(\d+)', description)
    if length_match:
        dimensions['length'] = length_match.group(1)

    # Извлекаем ширину лейки
    width_match = re.search(r'Ширина лейки:\s*(\d+)', description)
    if width_match:
        dimensions['width'] = width_match.group(1)

    # Извлекаем высоту держателя
    height_match = re.search(r'Высота держателя:\s*(\d+)', description)
    if height_match:
        dimensions['height'] = height_match.group(1)

    # Извлекаем размер упаковки
    package_match = re.search(r'Размер упаковки[:\s]*(\d+)[х×](\d+)[х×](\d+)', description)
    if package_match:
        dimensions['package_length'] = package_match.group(1)
        dimensions['package_width'] = package_match.group(2)
        dimensions['package_height'] = package_match.group(3)

    # Извлекаем вес
    weight_match = re.search(r'Вес товара:\s*(\d+)', description)
    if weight_match:
        dimensions['package_weight'] = weight_match.group(1)

    return dimensions

# Функция для извлечения цвета
def extract_color(description):
    if pd.isna(description):
        return None

    description = str(description).lower()

    colors = {
        'зеленый': 'зеленый',
        'синий': 'синий',
        'белый': 'белый',
        'черный': 'черный',
        'оранжевый': 'оранжевый'
    }

    for color_key, color_value in colors.items():
        if color_key in description:
            return color_value

    return None

# Функция для определения материала
def extract_material(description):
    if pd.isna(description):
        return None

    description = str(description).lower()

    if 'пластик' in description:
        return 'Пластик'
    elif 'металл' in description:
        return 'Металл'
    elif 'резина' in description:
        return 'Резина'

    return None

# Начинаем заполнение со строки 3 (первые две строки - заголовки)
row_num = 3

# Обрабатываем данные из products (начиная со строки 3, где данные о товарах)
for idx in range(3, len(products_df)):
    product_row = products_df.iloc[idx]

    # Пропускаем пустые строки
    if pd.isna(product_row[2]):
        continue

    # Колонка A (1) - Артикул
    if pd.notna(product_row[5]):
        sheet1.cell(row=row_num, column=1, value=str(product_row[5]))

    # Колонка B (2) - Наименование товара
    if pd.notna(product_row[2]):
        sheet1.cell(row=row_num, column=2, value=str(product_row[2]))

    # Колонка C (3) - Название модели
    if pd.notna(product_row[3]):
        sheet1.cell(row=row_num, column=3, value=str(product_row[3]))

    # Колонка D (4) - Характеристики товара (описание)
    if pd.notna(product_row[4]):
        sheet1.cell(row=row_num, column=4, value=str(product_row[4]))

    # Извлекаем размеры и другие параметры из описания
    dims = extract_dimensions(product_row[4])

    # Колонка J (10) - Длина
    if dims['length']:
        sheet1.cell(row=row_num, column=10, value=dims['length'])

    # Колонка K (11) - Ширина
    if dims['width']:
        sheet1.cell(row=row_num, column=11, value=dims['width'])

    # Колонка L (12) - Высота
    if dims['height']:
        sheet1.cell(row=row_num, column=12, value=dims['height'])

    # Колонка M (13) - Вес
    if dims['package_weight']:
        sheet1.cell(row=row_num, column=13, value=dims['package_weight'])

    # Колонка N (14) - Длина упаковки
    if dims['package_length']:
        sheet1.cell(row=row_num, column=14, value=dims['package_length'])

    # Колонка O (15) - Ширина упаковки
    if dims['package_width']:
        sheet1.cell(row=row_num, column=15, value=dims['package_width'])

    # Колонка P (16) - Высота упаковки
    if dims['package_height']:
        sheet1.cell(row=row_num, column=16, value=dims['package_height'])

    # Колонка Q (17) - Вес упаковки (используем тот же вес)
    if dims['package_weight']:
        sheet1.cell(row=row_num, column=17, value=dims['package_weight'])

    # Колонка E (5) - Страна производства (по умолчанию Россия)
    sheet1.cell(row=row_num, column=5, value="Россия")

    # Колонка I (9) - Единица измерения (по умолчанию шт)
    sheet1.cell(row=row_num, column=9, value="шт")

    # Колонка R (18) - Поливать дома или на улице (да/нет)
    sheet1.cell(row=row_num, column=18, value="да")

    # Колонка S (19) - Тип изделия
    if 'лейка' in str(product_row[2]).lower():
        sheet1.cell(row=row_num, column=19, value="Лейка")
    elif 'душ' in str(product_row[2]).lower():
        sheet1.cell(row=row_num, column=19, value="Душ садовый")
    else:
        sheet1.cell(row=row_num, column=19, value="Поливочное устройство")

    # Колонка T (20) - Цвет
    color = extract_color(product_row[4])
    if color:
        sheet1.cell(row=row_num, column=20, value=color)

    row_num += 1

# Сохраняем заполненный файл
template_wb.save('template_filled.xlsx')

print(f"Данные успешно перенесены! Заполнено {row_num - 3} строк")
print("Результат сохранен в файл: template_filled.xlsx")