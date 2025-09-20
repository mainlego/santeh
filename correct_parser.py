import pandas as pd
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Читаем исходный файл
products_df = pd.read_excel('products.xlsx', header=None)

def parse_product_description(description):
    """Правильно парсит описание товара и извлекает все параметры"""
    params = {
        'product_length': None,      # Длина изделия
        'product_width': None,       # Ширина изделия
        'product_height': None,      # Высота изделия
        'product_weight': None,      # Вес изделия
        'package_length': None,      # Длина упаковки
        'package_width': None,       # Ширина упаковки
        'package_height': None,      # Высота упаковки
        'package_weight': None,      # Вес упаковки
        'material': None,
        'coating': None,
        'country': None,
        'connection_size': None,
        'type': None,
        'color': None
    }

    if pd.isna(description):
        return params

    desc = str(description)

    # Извлекаем длину излива/трубки (это длина изделия)
    length_patterns = [
        r'Длина излива:\s*(\d+)\s*мм',
        r'Длина трубки:\s*(\d+)\s*мм',
        r'Длина:\s*(\d+)\s*мм'
    ]
    for pattern in length_patterns:
        match = re.search(pattern, desc)
        if match:
            params['product_length'] = int(match.group(1))
            break

    # Извлекаем ширину излива/лейки (это ширина изделия)
    width_patterns = [
        r'Ширина излива:\s*(\d+)\s*мм',
        r'Ширина лейки:\s*(\d+)\s*мм',
        r'Ширина:\s*(\d+)\s*мм'
    ]
    for pattern in width_patterns:
        match = re.search(pattern, desc)
        if match:
            params['product_width'] = int(match.group(1))
            break

    # Извлекаем высоту излива/смесителя (это высота изделия)
    height_patterns = [
        r'Высота излива:\s*(\d+)\s*мм',
        r'Высота смесителя:\s*(\d+)\s*мм',
        r'Высота держателя:\s*(\d+)\s*мм',
        r'Высота:\s*(\d+)\s*мм'
    ]
    for pattern in height_patterns:
        match = re.search(pattern, desc)
        if match:
            params['product_height'] = int(match.group(1))
            break

    # Извлекаем размеры упаковки (ДхШхВ)
    package_patterns = [
        r'ДхШхВ упаковки\s*(\d+)х(\d+)х(\d+)\s*мм',
        r'Размер упаковки[:\s]*(\d+)[х×](\d+)[х×](\d+)',
        r'Габариты упаковки[:\s]*(\d+)[х×](\d+)[х×](\d+)'
    ]
    for pattern in package_patterns:
        match = re.search(pattern, desc)
        if match:
            params['package_length'] = int(match.group(1))
            params['package_width'] = int(match.group(2))
            params['package_height'] = int(match.group(3))
            break

    # Извлекаем вес товара/брутто
    weight_patterns = [
        r'Вес брутто:\s*(\d+)\s*г',
        r'Вес товара:\s*(\d+)\s*г',
        r'Вес:\s*(\d+)\s*г'
    ]
    for pattern in weight_patterns:
        match = re.search(pattern, desc)
        if match:
            weight_g = int(match.group(1))
            params['package_weight'] = round(weight_g / 1000, 3)  # переводим в кг
            params['product_weight'] = round(weight_g / 1000, 3)  # пока ставим тот же вес
            break

    # Извлекаем материал
    if 'пластик' in desc.lower():
        params['material'] = 'Пластик'
    elif 'металл' in desc.lower():
        params['material'] = 'Металл'
    elif 'латунь' in desc.lower():
        params['material'] = 'Латунь'
    elif 'нержавеющая сталь' in desc.lower():
        params['material'] = 'Нержавеющая сталь'

    # Извлекаем покрытие
    coating_patterns = [
        r'Покрытие:\s*([^\n]+)',
        r'Отделка:\s*([^\n]+)'
    ]
    for pattern in coating_patterns:
        match = re.search(pattern, desc)
        if match:
            params['coating'] = match.group(1).strip()
            break

    # Извлекаем страну производителя
    country_match = re.search(r'Страна производитель:\s*([^\n]+)', desc)
    if country_match:
        params['country'] = country_match.group(1).strip()

    # Извлекаем присоединительный размер
    connection_match = re.search(r'Присоединительный размер:\s*([^\n]+)', desc)
    if connection_match:
        params['connection_size'] = connection_match.group(1).strip()

    # Извлекаем тип излива
    type_match = re.search(r'Тип излива:\s*([^\n]+)', desc)
    if type_match:
        params['type'] = type_match.group(1).strip()

    return params

def generate_smart_description(product_name, model, params, original_desc):
    """Генерирует описание на основе характеристик товара"""

    name = str(product_name).lower()

    # Базовое описание в зависимости от типа товара
    if 'смеситель' in name and 'раковин' in name:
        base_desc = "Смеситель для раковины с надежным механизмом и долговечной конструкцией."

        # Добавляем детали на основе характеристик
        details = []

        if params['coating']:
            if 'хром' in params['coating'].lower():
                details.append("Стильное хромированное покрытие устойчиво к коррозии и легко очищается")
            elif 'никель' in params['coating'].lower():
                details.append("Матовое никелевое покрытие придает современный вид")

        if params['type']:
            if 'фиксированный' in params['type'].lower():
                details.append("Фиксированный излив обеспечивает стабильную подачу воды")
            elif 'поворотный' in params['type'].lower():
                details.append("Поворотный излив увеличивает функциональность")

        if params['connection_size']:
            details.append(f"Стандартное подключение {params['connection_size']}")

        if params['product_length']:
            details.append(f"Длина излива {params['product_length']} мм для удобного использования")

        if len(details) > 0:
            base_desc += " " + ". ".join(details) + "."

    elif 'распылитель' in name:
        if 'веерный' in name:
            base_desc = "Веерный распылитель создает равномерный поток воды для деликатного полива растений."
        elif 'импульсный' in name:
            base_desc = "Импульсный распылитель обеспечивает дальний полив больших площадей."
        else:
            base_desc = "Распылитель для эффективного полива садовых растений."

        details = []
        if params['material']:
            details.append(f"Корпус из прочного {params['material'].lower()}а")
        if params['connection_size']:
            details.append(f"Присоединение {params['connection_size']}")

        if len(details) > 0:
            base_desc += " " + ". ".join(details) + "."

    elif 'дождеватель' in name:
        base_desc = "Дождеватель для автоматического орошения участка."
        if 'круговой' in name:
            base_desc += " Круговое распыление воды обеспечивает равномерный полив."
        elif 'осциллирующий' in name:
            base_desc += " Осциллирующий механизм покрывает прямоугольную площадь."

    elif 'насадка' in name or 'пистолет' in name:
        base_desc = "Поливочная насадка с регулировкой режимов подачи воды."
        details = []
        if 'многофункциональная' in name:
            details.append("Множество режимов полива от струи до распыления")
        if params['material']:
            details.append(f"Эргономичная рукоятка из {params['material'].lower()}а")

        if len(details) > 0:
            base_desc += " " + ". ".join(details) + "."

    else:
        # Универсальное описание
        base_desc = f"Качественное поливочное оборудование для садово-огородных работ."
        if params['material']:
            base_desc += f" Изготовлено из {params['material'].lower()}а."

    return base_desc

def generate_additional_description(params):
    """Генерирует дополнительное описание на основе технических характеристик"""
    details = []

    if params['coating']:
        details.append(f"Покрытие: {params['coating']}")

    if params['country']:
        details.append(f"Производство: {params['country']}")

    if params['product_length'] and params['product_height']:
        details.append(f"Компактные размеры: {params['product_length']}×{params['product_height']} мм")
    elif params['product_length']:
        details.append(f"Длина изделия: {params['product_length']} мм")

    if params['material']:
        details.append(f"Материал: {params['material']}")

    if len(details) == 0:
        return "Соответствует всем стандартам качества и безопасности."

    return ". ".join(details) + "."

# Обрабатываем все товары
processed_products = []

for idx in range(3, len(products_df)):
    product_row = products_df.iloc[idx]

    if pd.isna(product_row[2]):
        continue

    # Извлекаем параметры
    params = parse_product_description(product_row[4])

    product = {
        'name': str(product_row[2]) if pd.notna(product_row[2]) else '',
        'model': str(product_row[3]) if pd.notna(product_row[3]) else '',
        'article': str(product_row[5]) if pd.notna(product_row[5]) else '',
        'original_description': str(product_row[4]) if pd.notna(product_row[4]) else '',
        'parameters': params,
        'description': generate_smart_description(product_row[2], product_row[3], params, product_row[4]),
        'additional_description': generate_additional_description(params)
    }

    processed_products.append(product)

print(f"Обработано {len(processed_products)} товаров")

# Сохраняем в JSON для проверки
with open('processed_products.json', 'w', encoding='utf-8') as f:
    json.dump(processed_products, f, ensure_ascii=False, indent=2)

# Создаем Excel по новому шаблону
wb = Workbook()
ws = wb.active
ws.title = "Товары для Петрович"

# Заголовки
headers = [
    "Наименование товара от поставщика",
    "Описание товара",
    "Дополнительное описание (Необязательное)",
    "Страна происхождения",
    "Код альтозиции «Петрович»",
    "Преимущество перед аналогом",
    "Ссылка на товар на вашем сайте",
    "Базовая единица измерения",
    "Длина, мм.",
    "Ширина, мм.",
    "Высота, мм",
    "Вес, кг",
    "Длина изделия, мм.",
    "Ширина изделия, мм.",
    "Высота изделия, мм",
    "Вес изделия, кг",
    "Контроль цены для покупателя (да/нет)",
    "Вид упаковки",
    "Штрих код",
    "Тип штрих кода"
]

# Заполняем заголовки
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Заполняем данные
for row_idx, product in enumerate(processed_products, 2):
    params = product['parameters']

    ws.cell(row=row_idx, column=1, value=product['name'])                        # A - Наименование
    ws.cell(row=row_idx, column=2, value=product['description'])                 # B - Описание
    ws.cell(row=row_idx, column=3, value=product['additional_description'])      # C - Доп описание
    ws.cell(row=row_idx, column=4, value=params.get('country', 'Россия'))       # D - Страна
    ws.cell(row=row_idx, column=5, value=product['article'])                     # E - Артикул
    ws.cell(row=row_idx, column=6, value="Высокое качество, надежность, оптимальная цена")  # F - Преимущества
    ws.cell(row=row_idx, column=7, value="")                                     # G - Ссылка
    ws.cell(row=row_idx, column=8, value="шт")                                   # H - Ед. измерения

    # Размеры и вес УПАКОВКИ
    ws.cell(row=row_idx, column=9, value=params['package_length'])               # I - Длина упаковки
    ws.cell(row=row_idx, column=10, value=params['package_width'])               # J - Ширина упаковки
    ws.cell(row=row_idx, column=11, value=params['package_height'])              # K - Высота упаковки
    ws.cell(row=row_idx, column=12, value=params['package_weight'])              # L - Вес упаковки

    # Размеры и вес ИЗДЕЛИЯ
    ws.cell(row=row_idx, column=13, value=params['product_length'])              # M - Длина изделия
    ws.cell(row=row_idx, column=14, value=params['product_width'])               # N - Ширина изделия
    ws.cell(row=row_idx, column=15, value=params['product_height'])              # O - Высота изделия
    ws.cell(row=row_idx, column=16, value=params['product_weight'])              # P - Вес изделия

    ws.cell(row=row_idx, column=17, value="да")                                  # Q - Контроль цены
    ws.cell(row=row_idx, column=18, value="Картонная коробка")                   # R - Вид упаковки
    ws.cell(row=row_idx, column=19, value="")                                    # S - Штрих код
    ws.cell(row=row_idx, column=20, value="EAN13")                               # T - Тип штрих кода

# Автоподбор ширины колонок
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save('template_petrovich_corrected.xlsx')

print("\nФайлы созданы:")
print("- processed_products.json - JSON с обработанными данными")
print("- template_petrovich_corrected.xlsx - исправленный Excel")

# Показываем примеры извлеченных данных
print("\n=== ПРИМЕРЫ ИЗВЛЕЧЕННЫХ ДАННЫХ ===")
for i, product in enumerate(processed_products[:3]):
    print(f"\nТовар {i+1}: {product['name']}")
    print(f"Описание: {product['description']}")
    print(f"Параметры изделия: Д={product['parameters']['product_length']}мм, Ш={product['parameters']['product_width']}мм, В={product['parameters']['product_height']}мм")
    print(f"Параметры упаковки: {product['parameters']['package_length']}x{product['parameters']['package_width']}x{product['parameters']['package_height']}мм, {product['parameters']['package_weight']}кг")