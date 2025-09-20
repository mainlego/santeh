import pandas as pd
import json
import re
from openpyxl import Workbook

# Читаем исходный файл
products_df = pd.read_excel('products.xlsx', header=None)

def clean_text(text):
    """Очищает текст от недопустимых символов"""
    if pd.isna(text):
        return ""

    text = str(text)

    # Убираем недопустимые спецсимволы
    forbidden_chars = '!@#$%^&*()|?'
    for char in forbidden_chars:
        text = text.replace(char, '')

    # Убираем лишние пробелы
    text = re.sub(r'\s+', ' ', text).strip()

    # Ограничиваем длину до 8000 символов
    if len(text) > 8000:
        text = text[:8000]

    return text

def clean_number(value):
    """Очищает и проверяет числовые значения"""
    if pd.isna(value) or value is None:
        return None

    # Если это строка с цифрами, убираем пробелы
    if isinstance(value, str):
        value = value.replace(' ', '')
        try:
            value = float(value)
        except:
            return None

    # Проверяем, что это число
    if not isinstance(value, (int, float)):
        return None

    # Проверяем на отрицательные значения
    if value < 0:
        return None

    # Ограничиваем до 8 знаков до запятой и 2 после
    if value >= 100000000:  # 8 знаков до запятой
        return None

    # Округляем до 2 знаков после запятой
    return round(float(value), 2)

def ultimate_parse_description(description):
    """Парсер с очисткой данных"""
    params = {
        'product_length': None,
        'product_width': None,
        'product_height': None,
        'product_weight': None,
        'package_length': None,
        'package_width': None,
        'package_height': None,
        'package_weight': None,
        'material': None,
        'coating': None,
        'country': None,
        'connection_size': None,
        'type': None,
        'mount_type': None,
        'hose_type': None,
        'pressure_min': None,
        'pressure_max': None,
        'temperature_min': None,
        'temperature_max': None,
        'voltage_main': None,
        'voltage_backup': None,
        'zone_min': None,
        'zone_max': None,
        'delay': None,
        'max_pressure_bar': None,
        'cartridge_size': None,
        'hose_length': None
    }

    if pd.isna(description):
        return params

    desc = str(description)

    # 1. ДЛИНА ИЗДЕЛИЯ
    length_patterns = [
        r'Длина излива:\s*(\d+)\s*мм',
        r'Длина трубки:\s*(\d+)\s*мм',
        r'Длина:\s*(\d+)\s*мм',
        r'Длина ручки:\s*(\d+)\s*мм',
        r'Длина рукоятки:\s*(\d+)\s*мм',
        r'Длина общая:\s*(\d+)\s*мм'
    ]
    for pattern in length_patterns:
        match = re.search(pattern, desc)
        if match:
            params['product_length'] = clean_number(int(match.group(1)))
            break

    # Специальные форматы длины (в скобках)
    special_length = re.search(r'Длина ручки[:\s]*(\d+)\s*мм[^)]*\((\d+)\s*мм\)', desc)
    if special_length and not params['product_length']:
        params['product_length'] = clean_number(int(special_length.group(2)))

    # 2. ШИРИНА ИЗДЕЛИЯ
    width_patterns = [
        r'Ширина лейки:\s*(\d+)\s*мм',
        r'Ширина излива:\s*(\d+)\s*мм',
        r'Ширина:\s*(\d+)\s*мм',
        r'Ширина ручки:\s*(\d+)\s*мм'
    ]
    for pattern in width_patterns:
        match = re.search(pattern, desc)
        if match:
            params['product_width'] = clean_number(int(match.group(1)))
            break

    # Специальные форматы ширины
    special_width = re.search(r'Ширина ручки[:\s]*(\d+)\s*мм[^)]*\((\d+)\s*мм\)', desc)
    if special_width and not params['product_width']:
        params['product_width'] = clean_number(int(special_width.group(2)))

    # 3. ВЫСОТА ИЗДЕЛИЯ
    height_patterns = [
        r'Высота общая:\s*(\d+)\s*мм',
        r'Высота смесителя:\s*(\d+)\s*мм',
        r'Высота излива:\s*(\d+)\s*мм',
        r'Высота держателя:\s*(\d+)\s*мм',
        r'Высота:\s*(\d+)\s*мм',
        r'Высота ручки:\s*(\d+)\s*мм'
    ]
    for pattern in height_patterns:
        match = re.search(pattern, desc)
        if match:
            params['product_height'] = clean_number(int(match.group(1)))
            break

    # 4. РАЗМЕРЫ УПАКОВКИ
    package_patterns = [
        r'ДхШхВ упаковки\s*(\d+)[х×](\d+)[х×](\d+)\s*мм',
        r'Размер упаковки[:\s]*(\d+)[х×](\d+)[х×](\d+)',
        r'Габариты упаковки[:\s]*(\d+)[х×](\d+)[х×](\d+)',
        r'Упаковка[:\s]*(\d+)[х×](\d+)[х×](\d+)',
        r'(\d+)[х×](\d+)[х×](\d+)\s*мм'
    ]

    for pattern in package_patterns:
        matches = re.findall(pattern, desc)
        if matches:
            match = matches[0]
            params['package_length'] = clean_number(int(match[0]))
            params['package_width'] = clean_number(int(match[1]))
            params['package_height'] = clean_number(int(match[2]))
            break

    # 5. ВЕС
    weight_patterns = [
        r'Вес брутто:\s*(\d+)\s*г',
        r'Вес товара:\s*(\d+)\s*г',
        r'Вес:\s*(\d+)\s*г',
        r'(\d+)\s*560\s*г'
    ]
    for pattern in weight_patterns:
        match = re.search(pattern, desc)
        if match:
            weight_g = int(match.group(1))
            weight_kg = clean_number(weight_g / 1000)
            params['package_weight'] = weight_kg
            params['product_weight'] = weight_kg
            break

    # Специальный парсинг веса из конца строки
    weight_end = re.search(r'(\d+)\s+(\d+)\s*г$', desc)
    if weight_end and not params['package_weight']:
        combined = weight_end.group(1) + weight_end.group(2)
        weight_g = int(combined)
        weight_kg = clean_number(weight_g / 1000)
        params['package_weight'] = weight_kg
        params['product_weight'] = weight_kg

    # 6. ПОКРЫТИЕ
    coating_patterns = [
        r'Покрытие:\s*([^\n]+)',
        r'Отделка:\s*([^\n]+)',
        r'Материал покрытия:\s*([^\n]+)'
    ]
    for pattern in coating_patterns:
        match = re.search(pattern, desc)
        if match:
            params['coating'] = clean_text(match.group(1))
            break

    # 7. СТРАНА ПРОИЗВОДИТЕЛЬ
    country_match = re.search(r'Страна производитель:\s*([^\n]+)', desc)
    if country_match:
        params['country'] = clean_text(country_match.group(1))

    # 8. ПРИСОЕДИНИТЕЛЬНЫЙ РАЗМЕР
    connection_patterns = [
        r'Присоединительный размер:\s*([^\n]+)',
        r'Подключение:\s*([^\n]+)',
        r'Резьба:\s*([^\n]+)'
    ]
    for pattern in connection_patterns:
        match = re.search(pattern, desc)
        if match:
            params['connection_size'] = clean_text(match.group(1))
            break

    # 9. ТИП ИЗЛИВА/УСТРОЙСТВА
    type_patterns = [
        r'Тип излива:\s*([^\n]+)',
        r'Тип устройства:\s*([^\n]+)',
        r'Тип:\s*([^\n]+)'
    ]
    for pattern in type_patterns:
        match = re.search(pattern, desc)
        if match:
            params['type'] = clean_text(match.group(1))
            break

    # 10. МОНТАЖ
    mount_match = re.search(r'Монтаж:\s*([^\n]+)', desc)
    if mount_match:
        params['mount_type'] = clean_text(mount_match.group(1))

    # 11. ПОДВОДКА
    hose_match = re.search(r'Подводка:\s*([^\n]+)', desc)
    if hose_match:
        params['hose_type'] = clean_text(hose_match.group(1))

    # 12. ДАВЛЕНИЕ
    pressure_match = re.search(r'(\d+(?:[.,]\d+)?)\s*-\s*(\d+(?:[.,]\d+)?)\s*[Мм][Пп][Аа]', desc)
    if pressure_match:
        params['pressure_min'] = clean_number(float(pressure_match.group(1).replace(',', '.')))
        params['pressure_max'] = clean_number(float(pressure_match.group(2).replace(',', '.')))

    # 13. ТЕМПЕРАТУРА
    temp_match = re.search(r'([+-]?\d+)\s*[°C°С]\s*до\s*([+-]?\d+)\s*[°C°С]', desc)
    if temp_match:
        params['temperature_min'] = clean_number(int(temp_match.group(1)))
        params['temperature_max'] = clean_number(int(temp_match.group(2)))

    # 14. МАТЕРИАЛ
    material_patterns = [
        r'Материал:\s*([^\n]+)',
        r'Изготовлен из:\s*([^\n]+)'
    ]
    for pattern in material_patterns:
        match = re.search(pattern, desc)
        if match:
            params['material'] = clean_text(match.group(1))
            break

    # СПЕЦИАЛЬНЫЕ ПАРАМЕТРЫ ДЛЯ СЕНСОРНЫХ СМЕСИТЕЛЕЙ

    # Длина излива для сенсорных смесителей
    sensor_length_match = re.search(r'Длина излива[,\s]*мм:\s*(\d+)', desc)
    if sensor_length_match and not params['product_length']:
        params['product_length'] = clean_number(int(sensor_length_match.group(1)))

    # Высота излива для сенсорных смесителей
    sensor_height_match = re.search(r'Высота излива[,\s]*мм:\s*(\d+)', desc)
    if sensor_height_match and not params['product_height']:
        params['product_height'] = clean_number(int(sensor_height_match.group(1)))

    # Высота смесителя
    mixer_height_match = re.search(r'Высота смесителя:\s*(\d+)\s*мм', desc)
    if mixer_height_match and not params['product_height']:
        params['product_height'] = clean_number(int(mixer_height_match.group(1)))

    # Максимальное давление в барах
    max_pressure_match = re.search(r'Максимальное давление:\s*(\d+)\s*бар', desc)
    if max_pressure_match:
        params['max_pressure_bar'] = clean_number(int(max_pressure_match.group(1)))

    # Питание основное
    voltage_main = re.search(r'(\d+)\s*[Вв]ольт', desc)
    if voltage_main:
        params['voltage_main'] = clean_number(int(voltage_main.group(1)))

    # Резервное питание
    voltage_backup = re.search(r'резервное питание\s*(\d+)\s*[Вв]ольт', desc)
    if voltage_backup:
        params['voltage_backup'] = clean_number(int(voltage_backup.group(1)))

    # Зона срабатывания
    zone_match = re.search(r'(\d+)-(\d+)\s*см', desc)
    if zone_match:
        params['zone_min'] = clean_number(int(zone_match.group(1)))
        params['zone_max'] = clean_number(int(zone_match.group(2)))

    # Задержка срабатывания
    delay_match = re.search(r'(\d+[.,]?\d*)\s*сек', desc)
    if delay_match:
        params['delay'] = clean_number(float(delay_match.group(1).replace(',', '.')))

    # Керамический картридж
    cartridge_match = re.search(r'керамический\s*(\d+)\s*мм', desc)
    if cartridge_match:
        params['cartridge_size'] = clean_number(int(cartridge_match.group(1)))

    # Гибкая подводка
    hose_match = re.search(r'гибкая подводка\s*(\d+)\s*см', desc)
    if hose_match:
        params['hose_length'] = clean_number(int(hose_match.group(1)))

    # Дополнительные специальные размеры в конце описания
    end_dimensions = re.search(r'Длина ручки[,\s]*мм:\s*(\d+)[^.]*\.\s*Ширина ручки[,\s]*мм:\s*(\d+)', desc)
    if end_dimensions:
        if not params['product_length']:
            params['product_length'] = clean_number(int(end_dimensions.group(1)))
        if not params['product_width']:
            params['product_width'] = clean_number(int(end_dimensions.group(2)))

    return params

def create_clean_description(name, model, params, original_desc):
    """Создает чистое описание товара без спецсимволов"""
    name_lower = str(name).lower()
    model_str = str(model) if pd.notna(model) else ""

    # СМЕСИТЕЛИ
    if 'смеситель' in name_lower:
        # СЕНСОРНЫЕ СМЕСИТЕЛИ
        if 'сенсорный' in name_lower or 'sensor' in name_lower:
            desc = "Сенсорный смеситель с автоматическим включением воды. "

            # Добавляем технические характеристики
            if params['product_length']:
                desc += f"Длина излива {params['product_length']} мм. "
            if params['product_height']:
                desc += f"Высота излива {params['product_height']} мм. "

            if params['pressure_min'] and params['pressure_max']:
                desc += f"Рабочее давление {params['pressure_min']}-{params['pressure_max']} МПа. "

            if params['temperature_min'] and params['temperature_max']:
                desc += f"Температура воды от {params['temperature_min']} до {params['temperature_max']} градусов. "

            if params['voltage_main']:
                desc += f"Питание {params['voltage_main']} В"
                if params['voltage_backup']:
                    desc += f", резервное питание {params['voltage_backup']} В. "
                else:
                    desc += ". "

            if params['zone_min'] and params['zone_max']:
                desc += f"Зона срабатывания {params['zone_min']}-{params['zone_max']} см. "

            if params['delay']:
                desc += f"Задержка срабатывания {params['delay']} сек. "

            if params['max_pressure_bar']:
                desc += f"Максимальное давление {params['max_pressure_bar']} бар. "

        # ОБЫЧНЫЕ СМЕСИТЕЛИ
        else:
            if 'раковин' in name_lower:
                desc = "Смеситель для раковины"
            elif 'мойк' in name_lower:
                desc = "Смеситель для кухонной мойки"
            elif 'умывальник' in name_lower:
                desc = "Смеситель для умывальника"
            elif 'гигиенический' in name_lower:
                desc = "Гигиенический смеситель с лейкой"
            else:
                desc = "Смеситель"

            # Тип крепления
            if 'шпилька' in model_str.lower():
                desc += " с креплением на шпильку"
            elif 'гайка' in model_str.lower():
                desc += " с креплением на гайку"

            desc += ". "

            # Тип излива
            if params['type']:
                if 'фиксированный' in params['type'].lower():
                    desc += "Фиксированный излив обеспечивает стабильную подачу воды. "
                elif 'поворотный' in params['type'].lower():
                    desc += "Поворотный излив увеличивает удобство использования. "

            # Покрытие
            if params['coating']:
                if 'хром' in params['coating'].lower():
                    desc += "Хромированное покрытие устойчиво к коррозии. "
                elif 'матовый' in params['coating'].lower():
                    desc += "Матовое покрытие придает современный вид. "

            # Размеры
            if params['product_length']:
                desc += f"Длина излива {params['product_length']} мм. "

            # Давление
            if params['pressure_min'] and params['pressure_max']:
                desc += f"Рабочее давление {params['pressure_min']}-{params['pressure_max']} МПа. "

    # РАСПЫЛИТЕЛИ СПЕЦИАЛЬНЫЕ
    elif 'распылитель' in name_lower and 'автоматический' in name_lower:
        desc = "Автоматический распылитель для систем полива. "

        if 'таймер' in str(original_desc).lower():
            desc += "Встроенный таймер для программирования режимов полива. "

        if params['pressure_min'] and params['pressure_max']:
            desc += f"Рабочее давление {params['pressure_min']}-{params['pressure_max']} МПа. "

        if params['temperature_min'] and params['temperature_max']:
            desc += f"Рабочая температура от {params['temperature_min']} до {params['temperature_max']} градусов. "

    # РАСПЫЛИТЕЛИ ОБЫЧНЫЕ
    elif 'распылитель' in name_lower:
        if 'веерный' in name_lower:
            desc = "Веерный распылитель для равномерного полива. "
        elif 'импульсный' in name_lower:
            desc = "Импульсный распылитель для дальнего полива. "
        else:
            desc = "Распылитель для садового полива. "

    # ДОЖДЕВАТЕЛИ
    elif 'дождеватель' in name_lower:
        if 'круговой' in name_lower:
            desc = "Круговой дождеватель для орошения по окружности. "
        elif 'осциллирующий' in name_lower:
            desc = "Осциллирующий дождеватель для прямоугольных участков. "
        else:
            desc = "Дождеватель для автоматического полива. "

    # ПИСТОЛЕТЫ И НАСАДКИ
    elif 'пистолет' in name_lower or 'насадка' in name_lower:
        desc = "Поливочная насадка с регулировкой режимов. "
        if 'многофункциональная' in name_lower:
            desc += "Несколько режимов распыления. "

    # ШЛАНГИ
    elif 'шланг' in name_lower:
        desc = "Поливочный шланг для садовых работ. "
        if 'армированный' in name_lower:
            desc += "Армированная конструкция повышает прочность. "

    else:
        desc = "Качественное изделие для полива и водоснабжения. "

    # Общие характеристики
    if params['country']:
        desc += f"Производство: {params['country']}. "

    if params['connection_size']:
        desc += f"Присоединение: {params['connection_size']}. "

    return clean_text(desc)

# Обрабатываем ВСЕ товары
processed_products = []

print("=== СОЗДАНИЕ ЧИСТОГО EXCEL БЕЗ ФОРМАТИРОВАНИЯ ===")

for idx in range(3, len(products_df)):
    product_row = products_df.iloc[idx]

    if pd.isna(product_row[2]) or str(product_row[2]).strip() == '-':
        continue

    # Парсинг с очисткой
    params = ultimate_parse_description(product_row[4])

    name = clean_text(product_row[2]) if pd.notna(product_row[2]) else ''
    model = clean_text(product_row[3]) if pd.notna(product_row[3]) else ''
    article = clean_text(product_row[5]) if pd.notna(product_row[5]) else ''

    description = create_clean_description(name, model, params, product_row[4])

    # Дополнительное описание
    details = []
    if params['coating']:
        details.append(f"Покрытие: {params['coating']}")
    if params['mount_type']:
        details.append(f"Монтаж: {params['mount_type'].lower()}")
    if params['connection_size']:
        details.append(f"Соединение: {params['connection_size']}")
    if params['pressure_min'] and params['pressure_max']:
        details.append(f"Давление: {params['pressure_min']}-{params['pressure_max']} МПа")

    # Дополнительные детали для сенсорных смесителей
    if 'сенсорный' in name.lower() or 'sensor' in name.lower():
        if params['cartridge_size']:
            details.append(f"Керамический картридж: {params['cartridge_size']} мм")
        if params['hose_length']:
            details.append(f"Гибкая подводка: {params['hose_length']} см")

    additional_desc = clean_text(". ".join(details) + "." if details else "Соответствует стандартам качества и безопасности.")

    product = {
        'name': name,
        'model': model,
        'article': article,
        'original_description': clean_text(product_row[4]) if pd.notna(product_row[4]) else '',
        'parameters': params,
        'description': description,
        'additional_description': additional_desc
    }

    processed_products.append(product)

print(f"ОБРАБОТАНО ТОВАРОВ: {len(processed_products)}")

# Создаем ЧИСТЫЙ Excel БЕЗ ФОРМАТИРОВАНИЯ
wb = Workbook()
ws = wb.active
ws.title = "Товары"

headers = [
    "Код альтозиции «Петрович»",
    "Наименование товара от поставщика",
    "Описание товара",
    "Дополнительное описание (Необязательное)",
    "Страна происхождения",
    "Преимущество перед аналогом",
    "Ссылка на товар на вашем сайте",
    "Базовая единица измерения",
    "Длина, мм.",
    "Ширина, мм.",
    "Высота, мм",
    "Вес, кг",
    "Длина изделия, мм.",
    "Ширина изделия, мм.",
    "Высота изделия, мм",
    "Вес изделия, кг",
    "Контроль цены для покупателя (да/нет)",
    "Вид упаковки",
    "Штрих код",
    "Тип штрих кода"
]

# Заголовки БЕЗ ФОРМАТИРОВАНИЯ
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Данные БЕЗ ФОРМАТИРОВАНИЯ
for row_idx, product in enumerate(processed_products, 2):
    params = product['parameters']

    # Артикул (первая колонка)
    ws.cell(row=row_idx, column=1, value=product['article'])

    # Наименование
    ws.cell(row=row_idx, column=2, value=product['name'])

    # Описание
    ws.cell(row=row_idx, column=3, value=product['description'])

    # Дополнительное описание
    ws.cell(row=row_idx, column=4, value=product['additional_description'])

    # Страна
    ws.cell(row=row_idx, column=5, value="Россия")

    # Преимущества
    ws.cell(row=row_idx, column=6, value="Высокое качество; Надежность; Простота установки")

    # Пустые поля
    ws.cell(row=row_idx, column=7, value="")  # Ссылка
    ws.cell(row=row_idx, column=8, value="шт")  # Единица измерения

    # УПАКОВКА (9-12) - только числа без пробелов
    package_data = [params['package_length'], params['package_width'],
                   params['package_height'], params['package_weight']]

    for i, value in enumerate(package_data, 9):
        if value is not None:
            ws.cell(row=row_idx, column=i, value=value)
        else:
            ws.cell(row=row_idx, column=i, value="")

    # ИЗДЕЛИЕ (13-16) - только числа без пробелов
    product_data = [params['product_length'], params['product_width'],
                   params['product_height'], params['product_weight']]

    for i, value in enumerate(product_data, 13):
        if value is not None:
            ws.cell(row=row_idx, column=i, value=value)
        else:
            ws.cell(row=row_idx, column=i, value="")

    # Остальные поля
    ws.cell(row=row_idx, column=17, value="да")
    ws.cell(row=row_idx, column=18, value="Картонная коробка")
    ws.cell(row=row_idx, column=19, value="")  # Штрих код
    ws.cell(row=row_idx, column=20, value="EAN13")

# Информация о штрих-кодах
last_row = len(processed_products) + 3
ws.cell(row=last_row, column=1, value="ИНФОРМАЦИЯ О ШТРИХ-КОДАХ:")

last_row += 1
info_text = "Штрих-коды для всех товаров будут предоставлены дополнительно после согласования номенклатуры."
ws.cell(row=last_row, column=1, value=info_text)

# НЕ устанавливаем закрепление областей
# НЕ применяем форматирование

wb.save('template_petrovich_CLEAN.xlsx')

print("\nСОЗДАН ЧИСТЫЙ EXCEL!")
print("Файл: template_petrovich_CLEAN.xlsx")
print("+ Убраны все спецсимволы и форматирование")
print("+ Убраны пробелы между цифрами")
print("+ Ограничена длина описаний до 8000 символов")
print("+ Числовые поля проверены на корректность")
print("+ НЕТ закрепления областей и стилей")