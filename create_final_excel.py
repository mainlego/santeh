import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Читаем обработанные данные
with open('final_products.json', 'r', encoding='utf-8') as f:
    processed_products = json.load(f)

# Фильтруем товары - убираем товары с названием "-"
filtered_products = []
for product in processed_products:
    if product['name'].strip() != '-' and product['name'].strip() != '':
        filtered_products.append(product)

print(f"Отфильтрованы товары: было {len(processed_products)}, стало {len(filtered_products)}")

# Создаем новый Excel файл
wb = Workbook()
ws = wb.active
ws.title = "Товары для Петрович"

# Заголовки в правильном порядке (артикул первый)
headers = [
    "Код альтозиции «Петрович»",          # A - Артикул первый!
    "Наименование товара от поставщика",   # B
    "Описание товара",                     # C
    "Дополнительное описание (Необязательное)",  # D
    "Страна происхождения",                # E
    "Преимущество перед аналогом",         # F
    "Ссылка на товар на вашем сайте",      # G
    "Базовая единица измерения",           # H
    "Длина, мм.",                          # I - Упаковка
    "Ширина, мм.",                         # J - Упаковка
    "Высота, мм",                          # K - Упаковка
    "Вес, кг",                            # L - Упаковка
    "Длина изделия, мм.",                  # M - Изделие
    "Ширина изделия, мм.",                 # N - Изделие
    "Высота изделия, мм",                  # O - Изделие
    "Вес изделия, кг",                    # P - Изделие
    "Контроль цены для покупателя (да/нет)", # Q
    "Вид упаковки",                        # R
    "Штрих код",                          # S
    "Тип штрих кода"                      # T
]

# Стили для заголовков
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Стили для данных
data_font = Font(name='Calibri', size=10)
data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center')

# Стили для описаний
description_font = Font(name='Calibri', size=10)
description_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Границы
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Заполняем заголовки с красивым оформлением
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Устанавливаем высоту строки заголовка
ws.row_dimensions[1].height = 60

# Заполняем данные только для отфильтрованных товаров
for row_idx, product in enumerate(filtered_products, 2):
    params = product['parameters']

    # A - Артикул (первое поле!)
    cell_a = ws.cell(row=row_idx, column=1, value=product['article'])
    cell_a.font = Font(name='Calibri', size=10, bold=True)
    cell_a.alignment = center_alignment
    cell_a.border = thin_border

    # B - Наименование товара
    cell_b = ws.cell(row=row_idx, column=2, value=product['name'])
    cell_b.font = Font(name='Calibri', size=10, bold=True)
    cell_b.alignment = data_alignment
    cell_b.border = thin_border

    # C - Описание товара (увеличенная высота)
    cell_c = ws.cell(row=row_idx, column=3, value=product['description'])
    cell_c.font = description_font
    cell_c.alignment = description_alignment
    cell_c.border = thin_border

    # D - Дополнительное описание
    cell_d = ws.cell(row=row_idx, column=4, value=product['additional_description'])
    cell_d.font = description_font
    cell_d.alignment = description_alignment
    cell_d.border = thin_border

    # E - Страна происхождения (ВСЕГДА РОССИЯ!)
    cell_e = ws.cell(row=row_idx, column=5, value="Россия")
    cell_e.font = data_font
    cell_e.alignment = center_alignment
    cell_e.border = thin_border

    # F - Преимущества
    advantages = "Высокое качество материалов; Надежная конструкция; Простота установки"
    cell_f = ws.cell(row=row_idx, column=6, value=advantages)
    cell_f.font = data_font
    cell_f.alignment = data_alignment
    cell_f.border = thin_border

    # G - Ссылка на товар
    cell_g = ws.cell(row=row_idx, column=7, value="")
    cell_g.border = thin_border

    # H - Единица измерения
    cell_h = ws.cell(row=row_idx, column=8, value="шт")
    cell_h.font = data_font
    cell_h.alignment = center_alignment
    cell_h.border = thin_border

    # I-L - Размеры и вес УПАКОВКИ
    package_data = [
        params['package_length'],
        params['package_width'],
        params['package_height'],
        params['package_weight']
    ]

    for i, value in enumerate(package_data, 9):
        cell = ws.cell(row=row_idx, column=i, value=value)
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = thin_border
        if value:
            cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')

    # M-P - Размеры и вес ИЗДЕЛИЯ
    product_data = [
        params['product_length'],
        params['product_width'],
        params['product_height'],
        params['product_weight']
    ]

    for i, value in enumerate(product_data, 13):
        cell = ws.cell(row=row_idx, column=i, value=value)
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = thin_border
        if value:
            cell.fill = PatternFill(start_color='FFF2E7', end_color='FFF2E7', fill_type='solid')

    # Q - Контроль цены
    cell_q = ws.cell(row=row_idx, column=17, value="да")
    cell_q.font = data_font
    cell_q.alignment = center_alignment
    cell_q.border = thin_border

    # R - Вид упаковки
    cell_r = ws.cell(row=row_idx, column=18, value="Картонная коробка")
    cell_r.font = data_font
    cell_r.alignment = center_alignment
    cell_r.border = thin_border

    # S - Штрих код (пустой)
    cell_s = ws.cell(row=row_idx, column=19, value="")
    cell_s.border = thin_border

    # T - Тип штрих кода
    cell_t = ws.cell(row=row_idx, column=20, value="EAN13")
    cell_t.font = data_font
    cell_t.alignment = center_alignment
    cell_t.border = thin_border

    # Устанавливаем высоту строки для лучшего отображения описаний
    ws.row_dimensions[row_idx].height = 80

# Настройка ширины колонок
column_widths = {
    'A': 15,  # Артикул
    'B': 35,  # Наименование
    'C': 60,  # Описание
    'D': 40,  # Дополнительное описание
    'E': 12,  # Страна
    'F': 45,  # Преимущества
    'G': 12,  # Ссылка
    'H': 8,   # Единица
    'I': 10,  # Длина упаковки
    'J': 10,  # Ширина упаковки
    'K': 10,  # Высота упаковки
    'L': 8,   # Вес упаковки
    'M': 12,  # Длина изделия
    'N': 12,  # Ширина изделия
    'O': 12,  # Высота изделия
    'P': 10,  # Вес изделия
    'Q': 12,  # Контроль цены
    'R': 15,  # Вид упаковки
    'S': 12,  # Штрих код
    'T': 12   # Тип штрих кода
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Добавляем заголовки для групп колонок
ws.merge_cells('I1:L1')
merged_cell = ws['I1']
merged_cell.value = "УПАКОВКА (мм, кг)"
merged_cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
merged_cell.fill = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
merged_cell.alignment = Alignment(horizontal='center', vertical='center')
merged_cell.border = thin_border

ws.merge_cells('M1:P1')
merged_cell2 = ws['M1']
merged_cell2.value = "ИЗДЕЛИЕ (мм, кг)"
merged_cell2.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
merged_cell2.fill = PatternFill(start_color='D2691E', end_color='D2691E', fill_type='solid')
merged_cell2.alignment = Alignment(horizontal='center', vertical='center')
merged_cell2.border = thin_border

# Вставляем новую строку для подзаголовков
ws.insert_rows(2)

# Подзаголовки для упаковки
package_headers = ["Длина, мм", "Ширина, мм", "Высота, мм", "Вес, кг"]
for i, header in enumerate(package_headers, 9):
    cell = ws.cell(row=2, column=i, value=header)
    cell.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Подзаголовки для изделия
product_headers = ["Длина, мм", "Ширина, мм", "Высота, мм", "Вес, кг"]
for i, header in enumerate(product_headers, 13):
    cell = ws.cell(row=2, column=i, value=header)
    cell.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='D2691E', end_color='D2691E', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Высота строки подзаголовков
ws.row_dimensions[2].height = 25

# Обновляем высоту строки основных заголовков
ws.row_dimensions[1].height = 40

# Замораживаем первые две строки для удобства просмотра
ws.freeze_panes = 'A3'

# ДОБАВЛЯЕМ ИНФОРМАЦИЮ О ШТРИХ-КОДАХ ВНИЗУ
last_row = len(filtered_products) + 3  # +2 для заголовков +1 для пустой строки

# Пустая строка
last_row += 1

# Заголовок информационного блока
info_cell = ws.cell(row=last_row, column=1, value="ИНФОРМАЦИЯ О ШТРИХ-КОДАХ:")
info_cell.font = Font(name='Calibri', size=12, bold=True, color='FF0000')
info_cell.alignment = Alignment(horizontal='left', vertical='center')

last_row += 1

# Текст информации
info_text = "Штрих-коды для всех товаров будут предоставлены дополнительно после согласования номенклатуры."
info_detail = ws.cell(row=last_row, column=1, value=info_text)
info_detail.font = Font(name='Calibri', size=10, italic=True)
info_detail.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Объединяем ячейки для информации
ws.merge_cells(f'A{last_row}:T{last_row}')

last_row += 1

# Дополнительная информация
contact_text = "Для получения штрих-кодов обращайтесь к менеджеру проекта."
contact_detail = ws.cell(row=last_row, column=1, value=contact_text)
contact_detail.font = Font(name='Calibri', size=10, italic=True, color='666666')
contact_detail.alignment = Alignment(horizontal='left', vertical='center')

# Объединяем ячейки для контактной информации
ws.merge_cells(f'A{last_row}:T{last_row}')

# Сохраняем файл
wb.save('template_petrovich_FINAL_CORRECTED.xlsx')

print("СОЗДАН ФИНАЛЬНЫЙ ИСПРАВЛЕННЫЙ EXCEL!")
print("Файл: template_petrovich_FINAL_CORRECTED.xlsx")
print(f"\nВключено товаров: {len(filtered_products)}")
print("\nИсправления:")
print("+ Артикул на первом месте")
print("+ Страна везде установлена как 'Россия'")
print("+ Добавлена информация о штрих-кодах внизу")
print("+ Убраны товары с названием '-'")
print("+ Красивое оформление с группировкой колонок")