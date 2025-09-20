import pandas as pd
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Читаем исходный файл
products_df = pd.read_excel('products.xlsx', header=None)

def enhanced_parse_description(description):
    """Улучшенный парсер, извлекающий ВСЕ возможные параметры"""
    params = {
        'product_length': None,
        'product_width': None,
        'product_height': None,
        'product_weight': None,
        'package_length': None,
        'package_width': None,
        'package_height': None,
        'package_weight': None,
        'material': None,
        'coating': None,
        'country': None,
        'connection_size': None,
        'type': None,
        'mount_type': None,
        'hose_type': None
    }

    if pd.isna(description):
        return params

    desc = str(description)

    # 1. ДЛИНА ИЗДЕЛИЯ (излива/трубки)
    length_patterns = [
        r'Длина излива:\s*(\d+)\s*мм',
        r'Длина трубки:\s*(\d+)\s*мм',
        r'Длина:\s*(\d+)\s*мм'
    ]
    for pattern in length_patterns:
        match = re.search(pattern, desc)
        if match:
            params['product_length'] = int(match.group(1))
            break

    # 2. ШИРИНА ИЗДЕЛИЯ (лейки/излива)
    width_patterns = [
        r'Ширина лейки:\s*(\d+)\s*мм',
        r'Ширина излива:\s*(\d+)\s*мм',
        r'Ширина:\s*(\d+)\s*мм'
    ]
    for pattern in width_patterns:
        match = re.search(pattern, desc)
        if match:
            params['product_width'] = int(match.group(1))
            break

    # 3. ВЫСОТА ИЗДЕЛИЯ (приоритет: смеситель > излив > держатель)
    height_patterns = [
        r'Высота смесителя:\s*(\d+)\s*мм',
        r'Высота излива:\s*(\d+)\s*мм',
        r'Высота держателя:\s*(\d+)\s*мм',
        r'Высота:\s*(\d+)\s*мм'
    ]
    for pattern in height_patterns:
        match = re.search(pattern, desc)
        if match:
            params['product_height'] = int(match.group(1))
            break

    # 4. РАЗМЕРЫ УПАКОВКИ (все возможные форматы)
    package_patterns = [
        r'ДхШхВ упаковки\s*(\d+)х(\d+)х(\d+)\s*мм',
        r'Размер упаковки[:\s]*(\d+)[х×](\d+)[х×](\d+)',
        r'Габариты упаковки[:\s]*(\d+)[х×](\d+)[х×](\d+)',
        r'Упаковка[:\s]*(\d+)[х×](\d+)[х×](\d+)',
        r'(\d+)[х×](\d+)[х×](\d+)\s*мм'  # Общий формат размеров
    ]

    for pattern in package_patterns:
        matches = re.findall(pattern, desc)
        if matches:
            # Берем первое совпадение
            match = matches[0]
            params['package_length'] = int(match[0])
            params['package_width'] = int(match[1])
            params['package_height'] = int(match[2])
            break

    # 5. ВЕС (все форматы)
    weight_patterns = [
        r'Вес брутто:\s*(\d+)\s*г',
        r'Вес товара:\s*(\d+)\s*г',
        r'Вес:\s*(\d+)\s*г'
    ]
    for pattern in weight_patterns:
        match = re.search(pattern, desc)
        if match:
            weight_g = int(match.group(1))
            params['package_weight'] = round(weight_g / 1000, 3)
            params['product_weight'] = round(weight_g / 1000, 3)
            break

    # 6. ПОКРЫТИЕ
    coating_match = re.search(r'Покрытие:\s*([^\n]+)', desc)
    if coating_match:
        params['coating'] = coating_match.group(1).strip()

    # 7. СТРАНА
    country_match = re.search(r'Страна производитель:\s*([^\n]+)', desc)
    if country_match:
        params['country'] = country_match.group(1).strip()

    # 8. ПРИСОЕДИНИТЕЛЬНЫЙ РАЗМЕР
    connection_match = re.search(r'Присоединительный размер:\s*([^\n]+)', desc)
    if connection_match:
        params['connection_size'] = connection_match.group(1).strip()

    # 9. ТИП ИЗЛИВА
    type_match = re.search(r'Тип излива:\s*([^\n]+)', desc)
    if type_match:
        params['type'] = type_match.group(1).strip()

    # 10. МОНТАЖ
    mount_match = re.search(r'Монтаж:\s*([^\n]+)', desc)
    if mount_match:
        params['mount_type'] = mount_match.group(1).strip()

    # 11. ПОДВОДКА
    hose_match = re.search(r'Подводка:\s*([^\n]+)', desc)
    if hose_match:
        params['hose_type'] = hose_match.group(1).strip()

    return params

def create_accurate_description(name, model, params, original_desc):
    """Создает точное описание товара"""
    name_lower = str(name).lower()
    model_str = str(model) if pd.notna(model) else ""

    # СМЕСИТЕЛИ ДЛЯ РАКОВИНЫ
    if 'смеситель' in name_lower and 'раковин' in name_lower:
        desc = f"Смеситель для раковины"

        if 'шпилька' in model_str.lower():
            desc += " с креплением на шпильку"
        elif 'гайка' in model_str.lower():
            desc += " с креплением на гайку"

        desc += ". "

        if params['type']:
            if 'фиксированный' in params['type'].lower():
                desc += "Фиксированный излив не поворачивается, обеспечивает стабильную подачу воды. "
            elif 'поворотный' in params['type'].lower():
                desc += "Поворотный излив увеличивает радиус действия и удобство использования. "

        if params['coating'] and 'хром' in params['coating'].lower():
            desc += "Хромированное покрытие придает блеск и защищает от коррозии. "

        if params['product_length']:
            desc += f"Длина излива {params['product_length']} мм. "

        if params['connection_size']:
            desc += f"Стандартное подключение {params['connection_size']}. "

    # СМЕСИТЕЛИ ДЛЯ МОЙКИ
    elif 'смеситель' in name_lower and 'мойк' in name_lower:
        desc = f"Смеситель для кухонной мойки"

        if 'гайка' in model_str.lower():
            desc += " с креплением на гайку"
        elif 'шпилька' in model_str.lower():
            desc += " с креплением на шпильку"

        desc += ". "

        if params['type'] and 'поворотный' in params['type'].lower():
            desc += "Поворотный излив обеспечивает максимальное удобство при мытье посуды. "

        if params['product_length']:
            desc += f"Длина излива {params['product_length']} мм для удобного доступа к раковине. "

        if params['coating'] and 'хром' in params['coating'].lower():
            desc += "Хромированная поверхность легко очищается и не тускнеет. "

    # СМЕСИТЕЛИ ДЛЯ УМЫВАЛЬНИКА
    elif 'смеситель' in name_lower and 'умывальник' in name_lower:
        desc = f"Смеситель для умывальника"

        if 'шпилька' in model_str.lower():
            desc += " с креплением на шпильку"
        elif 'гайка' in model_str.lower():
            desc += " с креплением на гайку"

        desc += ". "

        if params['type'] and 'поворотный' in params['type'].lower():
            desc += "Поворотный механизм излива для максимального комфорта. "

        if params['product_length']:
            desc += f"Длина излива {params['product_length']} мм. "

    # РАСПЫЛИТЕЛИ
    elif 'распылитель' in name_lower:
        if 'веерный' in name_lower:
            desc = "Веерный распылитель создает широкий поток воды в форме веера для равномерного полива растений. "
        elif 'импульсный' in name_lower:
            desc = "Импульсный распылитель обеспечивает дальнобойный полив за счет пульсирующих струй воды. "
        else:
            desc = "Распылитель для садового полива. "

        if 'зеленый' in model_str.lower():
            desc += "Зеленый цвет корпуса. "
        elif 'синий' in model_str.lower():
            desc += "Синий цвет корпуса. "

    # ДОЖДЕВАТЕЛИ
    elif 'дождеватель' in name_lower:
        if 'круговой' in name_lower:
            desc = "Круговой дождеватель обеспечивает равномерное орошение по окружности. "
        elif 'осциллирующий' in name_lower:
            desc = "Осциллирующий дождеватель покрывает прямоугольную площадь маятниковыми движениями. "
        else:
            desc = "Дождеватель для автоматического полива участка. "

    # НАСАДКИ И ПИСТОЛЕТЫ
    elif 'насадка' in name_lower or 'пистолет' in name_lower:
        desc = "Поливочная насадка с регулировкой режимов распыления. "
        if 'многофункциональная' in name_lower:
            desc += "Несколько режимов полива от тонкой струи до мягкого распыления. "

    # ШЛАНГИ
    elif 'шланг' in name_lower:
        desc = "Поливочный шланг для садовых работ. "
        if 'армированный' in name_lower:
            desc += "Армированная конструкция обеспечивает повышенную прочность. "

    # СОЕДИНИТЕЛИ
    elif 'соединитель' in name_lower or 'переходник' in name_lower:
        desc = "Соединительный элемент для поливочных систем. "

    else:
        desc = f"Качественное изделие для водоснабжения и полива. "

    if params['country']:
        desc += f"Производство: {params['country']}. "

    return desc.strip()

# Обрабатываем все товары с улучшенным парсером
processed_products = []

print("=== ПОЛНАЯ ОБРАБОТКА С УЛУЧШЕННЫМ ПАРСЕРОМ ===")

for idx in range(3, len(products_df)):
    product_row = products_df.iloc[idx]

    if pd.isna(product_row[2]) or str(product_row[2]).strip() == '-':
        continue

    # Улучшенный парсинг
    params = enhanced_parse_description(product_row[4])

    name = str(product_row[2]) if pd.notna(product_row[2]) else ''
    model = str(product_row[3]) if pd.notna(product_row[3]) else ''

    description = create_accurate_description(name, model, params, product_row[4])

    # Дополнительное описание
    details = []
    if params['coating']:
        details.append(f"Покрытие: {params['coating']}")
    if params['mount_type']:
        details.append(f"Тип монтажа: {params['mount_type'].lower()}")
    if params['connection_size']:
        details.append(f"Присоединение: {params['connection_size']}")

    additional_desc = ". ".join(details) + "." if details else "Соответствует стандартам качества."

    product = {
        'name': name,
        'model': model,
        'article': str(product_row[5]) if pd.notna(product_row[5]) else '',
        'original_description': str(product_row[4]) if pd.notna(product_row[4]) else '',
        'parameters': params,
        'description': description,
        'additional_description': additional_desc
    }

    processed_products.append(product)

    # Показываем прогресс для товаров с размерами
    if any([params['product_length'], params['product_width'], params['product_height'],
           params['package_length'], params['package_width'], params['package_height']]):
        print(f"{len(processed_products):2d}. {name[:40]}")
        if params['product_length'] or params['product_width'] or params['product_height']:
            print(f"    Изделие: {params['product_length']}x{params['product_width']}x{params['product_height']}мм")
        if params['package_length'] or params['package_width'] or params['package_height']:
            print(f"    Упаковка: {params['package_length']}x{params['package_width']}x{params['package_height']}мм")
        if params['package_weight']:
            print(f"    Вес: {params['package_weight']}кг")

print(f"\nОБРАБОТАНО ТОВАРОВ: {len(processed_products)}")

# Создаем Excel с полными данными
wb = Workbook()
ws = wb.active
ws.title = "Товары для Петрович - ПОЛНЫЕ ДАННЫЕ"

headers = [
    "Код альтозиции «Петрович»",
    "Наименование товара от поставщика",
    "Описание товара",
    "Дополнительное описание (Необязательное)",
    "Страна происхождения",
    "Преимущество перед аналогом",
    "Ссылка на товар на вашем сайте",
    "Базовая единица измерения",
    "Длина, мм.",
    "Ширина, мм.",
    "Высота, мм",
    "Вес, кг",
    "Длина изделия, мм.",
    "Ширина изделия, мм.",
    "Высота изделия, мм",
    "Вес изделия, кг",
    "Контроль цены для покупателя (да/нет)",
    "Вид упаковки",
    "Штрих код",
    "Тип штрих кода"
]

# Стили
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(name='Calibri', size=10)
data_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Заголовки
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

ws.row_dimensions[1].height = 60

# Данные
for row_idx, product in enumerate(processed_products, 2):
    params = product['parameters']

    # A - Артикул
    cell_a = ws.cell(row=row_idx, column=1, value=product['article'])
    cell_a.font = Font(name='Calibri', size=10, bold=True)
    cell_a.alignment = center_alignment
    cell_a.border = thin_border

    # B - Наименование
    cell_b = ws.cell(row=row_idx, column=2, value=product['name'])
    cell_b.font = Font(name='Calibri', size=10, bold=True)
    cell_b.alignment = data_alignment
    cell_b.border = thin_border

    # C - Описание
    cell_c = ws.cell(row=row_idx, column=3, value=product['description'])
    cell_c.font = data_font
    cell_c.alignment = data_alignment
    cell_c.border = thin_border

    # D - Дополнительное описание
    cell_d = ws.cell(row=row_idx, column=4, value=product['additional_description'])
    cell_d.font = data_font
    cell_d.alignment = data_alignment
    cell_d.border = thin_border

    # E - Страна (всегда Россия)
    cell_e = ws.cell(row=row_idx, column=5, value="Россия")
    cell_e.font = data_font
    cell_e.alignment = center_alignment
    cell_e.border = thin_border

    # F - Преимущества
    cell_f = ws.cell(row=row_idx, column=6, value="Высокое качество материалов; Надежная конструкция; Простота установки")
    cell_f.font = data_font
    cell_f.alignment = data_alignment
    cell_f.border = thin_border

    # G-H
    ws.cell(row=row_idx, column=7, value="").border = thin_border
    ws.cell(row=row_idx, column=8, value="шт").font = data_font
    ws.cell(row=row_idx, column=8).alignment = center_alignment
    ws.cell(row=row_idx, column=8).border = thin_border

    # I-L - УПАКОВКА
    package_data = [
        params['package_length'],
        params['package_width'],
        params['package_height'],
        params['package_weight']
    ]

    for i, value in enumerate(package_data, 9):
        cell = ws.cell(row=row_idx, column=i, value=value)
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = thin_border
        if value:
            cell.fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')

    # M-P - ИЗДЕЛИЕ
    product_data = [
        params['product_length'],
        params['product_width'],
        params['product_height'],
        params['product_weight']
    ]

    for i, value in enumerate(product_data, 13):
        cell = ws.cell(row=row_idx, column=i, value=value)
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = thin_border
        if value:
            cell.fill = PatternFill(start_color='FFF2E7', end_color='FFF2E7', fill_type='solid')

    # Q-T
    ws.cell(row=row_idx, column=17, value="да").font = data_font
    ws.cell(row=row_idx, column=17).alignment = center_alignment
    ws.cell(row=row_idx, column=17).border = thin_border

    ws.cell(row=row_idx, column=18, value="Картонная коробка").font = data_font
    ws.cell(row=row_idx, column=18).alignment = center_alignment
    ws.cell(row=row_idx, column=18).border = thin_border

    ws.cell(row=row_idx, column=19, value="").border = thin_border

    ws.cell(row=row_idx, column=20, value="EAN13").font = data_font
    ws.cell(row=row_idx, column=20).alignment = center_alignment
    ws.cell(row=row_idx, column=20).border = thin_border

    ws.row_dimensions[row_idx].height = 80

# Ширина колонок
column_widths = {
    'A': 15, 'B': 35, 'C': 60, 'D': 40, 'E': 12, 'F': 45, 'G': 12, 'H': 8,
    'I': 10, 'J': 10, 'K': 10, 'L': 8, 'M': 12, 'N': 12, 'O': 12, 'P': 10,
    'Q': 12, 'R': 15, 'S': 12, 'T': 12
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

# Добавляем информацию о штрих-кодах
last_row = len(processed_products) + 3

ws.cell(row=last_row, column=1, value="ИНФОРМАЦИЯ О ШТРИХ-КОДАХ:").font = Font(name='Calibri', size=12, bold=True, color='FF0000')

last_row += 1
info_text = "Штрих-коды для всех товаров будут предоставлены дополнительно после согласования номенклатуры."
ws.cell(row=last_row, column=1, value=info_text).font = Font(name='Calibri', size=10, italic=True)
ws.merge_cells(f'A{last_row}:T{last_row}')

last_row += 1
contact_text = "Для получения штрих-кодов обращайтесь к менеджеру проекта."
ws.cell(row=last_row, column=1, value=contact_text).font = Font(name='Calibri', size=10, italic=True, color='666666')
ws.merge_cells(f'A{last_row}:T{last_row}')

# Замораживаем заголовки
ws.freeze_panes = 'A2'

wb.save('template_petrovich_COMPLETE.xlsx')

print("\nСОЗДАН EXCEL С ПОЛНЫМИ ДАННЫМИ!")
print("Файл: template_petrovich_COMPLETE.xlsx")
print("+ Все размеры извлечены из описаний")
print("+ Правильное разделение на упаковку и изделие")
print("+ Улучшенный парсер для всех форматов размеров")