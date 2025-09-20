import pandas as pd
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Читаем JSON с товарами
with open('products_data.json', 'r', encoding='utf-8') as f:
    products = json.load(f)

# Функция для более точного извлечения параметров
def parse_description(full_description):
    """Извлекает параметры из полного описания"""
    params = {
        'length': None,
        'width': None,
        'height': None,
        'weight': None,
        'package_length': None,
        'package_width': None,
        'package_height': None,
        'package_weight': None,
        'material': 'Пластик',
        'color': None,
        'connection_size': '1/2"'
    }

    if not full_description:
        return params

    desc = str(full_description)

    # Длина трубки
    length_match = re.search(r'Длина трубки:\s*(\d+)\s*мм', desc)
    if length_match:
        params['length'] = int(length_match.group(1))

    # Ширина лейки
    width_match = re.search(r'Ширина лейки:\s*(\d+)\s*мм', desc)
    if width_match:
        params['width'] = int(width_match.group(1))

    # Высота держателя
    height_match = re.search(r'Высота держателя:\s*(\d+)\s*мм', desc)
    if height_match:
        params['height'] = int(height_match.group(1))

    # Размеры упаковки
    package_match = re.search(r'Размер упаковки[:\s]*(\d+)[х×](\d+)[х×](\d+)', desc)
    if package_match:
        params['package_length'] = int(package_match.group(1))
        params['package_width'] = int(package_match.group(2))
        params['package_height'] = int(package_match.group(3))

    # Вес товара
    weight_match = re.search(r'Вес товара:\s*(\d+)\s*г', desc)
    if weight_match:
        params['package_weight'] = round(int(weight_match.group(1)) / 1000, 3)  # в кг
        params['weight'] = params['package_weight']

    # Материал
    if 'пластик' in desc.lower():
        params['material'] = 'Пластик'
    elif 'металл' in desc.lower():
        params['material'] = 'Металл'
    elif 'латунь' in desc.lower():
        params['material'] = 'Латунь'

    # Цвет
    colors = {
        'зеленый': 'Зеленый', 'зелёный': 'Зеленый',
        'синий': 'Синий', 'белый': 'Белый',
        'серый': 'Серый', 'оранжевый': 'Оранжевый'
    }
    for color_key, color_value in colors.items():
        if color_key in desc.lower():
            params['color'] = color_value
            break

    return params

# Функция генерации описания товара
def generate_description(product):
    """Генерирует описание товара"""
    name = product['name'].lower()
    model = product.get('model', '').lower()

    if 'распылитель' in name and 'веерный' in name:
        return "Веерный распылитель для равномерного полива газонов и цветников. Создает мягкий веерообразный поток воды."

    elif 'распылитель' in name and 'импульсный' in name:
        return "Импульсный распылитель с регулируемой дальностью полива. Обеспечивает круговое орошение больших площадей."

    elif 'распылитель' in name and 'дождь' in model:
        return "Дождевой распылитель для мягкого полива растений. Имитирует естественный дождь."

    elif 'насадка' in name:
        return "Многофункциональная насадка для полива с регулировкой напора. Различные режимы распыления."

    elif 'дождеватель' in name and 'круговой' in name:
        return "Круговой дождеватель для автоматического полива. Вращающаяся конструкция обеспечивает равномерное орошение."

    elif 'дождеватель' in name and 'осциллирующий' in name:
        return "Осциллирующий дождеватель для прямоугольных участков. Маятниковый механизм для равномерного покрытия."

    elif 'пистолет' in name:
        return "Поливочный пистолет с плавной регулировкой потока. Удобная прорезиненная рукоятка."

    else:
        return "Садовое поливочное устройство для эффективного орошения растений."

# Функция генерации дополнительного описания
def generate_additional_description(product, params):
    """Генерирует дополнительное описание"""
    features = []

    if params['material']:
        features.append(f"Изготовлен из прочного {params['material'].lower()}а")

    if params['color']:
        features.append(f"Цвет: {params['color'].lower()}")

    if params['connection_size']:
        features.append(f"Стандартное соединение {params['connection_size']}")

    if 'регулируемый' in product['name'].lower():
        features.append("Регулируемые параметры полива")

    if len(features) > 0:
        return ". ".join(features) + "."
    else:
        return "Надежное и долговечное решение для садового полива."

# Функция генерации преимуществ
def generate_advantages(product):
    """Генерирует преимущества перед аналогами"""
    name = product['name'].lower()

    advantages = []

    if 'веерный' in name:
        advantages.append("Равномерное распределение воды")

    if 'импульсный' in name:
        advantages.append("Большая дальность полива")

    if 'регулируемый' in name:
        advantages.append("Точная настройка режимов")

    if 'многофункциональный' in name or 'насадка' in name:
        advantages.append("Несколько режимов полива в одном устройстве")

    # Общие преимущества
    advantages.extend([
        "Простота использования",
        "Долговечность конструкции",
        "Оптимальное соотношение цена-качество"
    ])

    return "; ".join(advantages[:3])  # Берем первые 3 преимущества

# Создаем новый Excel файл
wb = Workbook()
ws = wb.active
ws.title = "Товары для Петрович"

# Заголовки колонок
headers = [
    "Наименование товара от поставщика",
    "Описание товара",
    "Дополнительное описание (Необязательное)",
    "Страна происхождения",
    "Код альтозиции «Петрович»",
    "Преимущество перед аналогом",
    "Ссылка на товар на вашем сайте",
    "Базовая единица измерения",
    "Длина, мм.",
    "Ширина, мм.",
    "Высота, мм",
    "Вес, кг",
    "Длина изделия, мм.",
    "Ширина изделия, мм.",
    "Высота изделия, мм",
    "Вес изделия, кг",
    "Контроль цены для покупателя (да/нет)",
    "Вид упаковки",
    "Штрих код",
    "Тип штрих кода"
]

# Заполняем заголовки
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    cell.alignment = Alignment(wrap_text=True, vertical="center")

# Заполняем данные товаров
for row_idx, product in enumerate(products, 2):
    params = parse_description(product['full_description'])

    # A - Наименование товара от поставщика
    ws.cell(row=row_idx, column=1, value=product['name'])

    # B - Описание товара
    ws.cell(row=row_idx, column=2, value=generate_description(product))

    # C - Дополнительное описание
    ws.cell(row=row_idx, column=3, value=generate_additional_description(product, params))

    # D - Страна происхождения
    ws.cell(row=row_idx, column=4, value="Россия")

    # E - Код альтозиции «Петрович» (артикул)
    ws.cell(row=row_idx, column=5, value=product['article'])

    # F - Преимущество перед аналогом
    ws.cell(row=row_idx, column=6, value=generate_advantages(product))

    # G - Ссылка на товар (пустая)
    ws.cell(row=row_idx, column=7, value="")

    # H - Базовая единица измерения
    ws.cell(row=row_idx, column=8, value="шт")

    # I - Длина упаковки, мм
    ws.cell(row=row_idx, column=9, value=params['package_length'])

    # J - Ширина упаковки, мм
    ws.cell(row=row_idx, column=10, value=params['package_width'])

    # K - Высота упаковки, мм
    ws.cell(row=row_idx, column=11, value=params['package_height'])

    # L - Вес упаковки, кг
    ws.cell(row=row_idx, column=12, value=params['package_weight'])

    # M - Длина изделия, мм
    ws.cell(row=row_idx, column=13, value=params['length'])

    # N - Ширина изделия, мм
    ws.cell(row=row_idx, column=14, value=params['width'])

    # O - Высота изделия, мм
    ws.cell(row=row_idx, column=15, value=params['height'])

    # P - Вес изделия, кг
    ws.cell(row=row_idx, column=16, value=params['weight'])

    # Q - Контроль цены для покупателя
    ws.cell(row=row_idx, column=17, value="да")

    # R - Вид упаковки
    ws.cell(row=row_idx, column=18, value="Картонная коробка")

    # S - Штрих код (пустой)
    ws.cell(row=row_idx, column=19, value="")

    # T - Тип штрих кода
    ws.cell(row=row_idx, column=20, value="EAN13")

# Автоматическая ширина колонок
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 50)
    ws.column_dimensions[column_letter].width = adjusted_width

# Сохраняем файл
wb.save('template_petrovich.xlsx')

print(f"Создан Excel файл для Петрович с {len(products)} товарами")
print("Файл сохранен как: template_petrovich.xlsx")
print("\nСтруктура файла:")
print("- Наименования товаров от поставщика")
print("- Описания товаров (сгенерированные)")
print("- Дополнительные описания")
print("- Извлеченные параметры (размеры, вес)")
print("- Преимущества перед аналогами")