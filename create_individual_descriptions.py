import pandas as pd
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Читаем исходный файл
products_df = pd.read_excel('products.xlsx', header=None)

def parse_product_description(description):
    """Извлекает параметры из описания"""
    params = {
        'product_length': None,
        'product_width': None,
        'product_height': None,
        'product_weight': None,
        'package_length': None,
        'package_width': None,
        'package_height': None,
        'package_weight': None,
        'material': None,
        'coating': None,
        'country': None,
        'connection_size': None,
        'type': None,
        'mount_type': None,
        'hose_type': None
    }

    if pd.isna(description):
        return params

    desc = str(description)

    # Длина излива/трубки
    length_match = re.search(r'Длина излива:\s*(\d+)\s*мм', desc)
    if length_match:
        params['product_length'] = int(length_match.group(1))

    # Ширина излива/лейки
    width_match = re.search(r'Ширина излива:\s*(\d+)\s*мм', desc)
    if width_match:
        params['product_width'] = int(width_match.group(1))

    # Высота излива/смесителя
    height_match = re.search(r'Высота излива:\s*(\d+)\s*мм', desc)
    if height_match:
        params['product_height'] = int(height_match.group(1))

    # Размеры упаковки
    package_match = re.search(r'ДхШхВ упаковки\s*(\d+)х(\d+)х(\d+)\s*мм', desc)
    if package_match:
        params['package_length'] = int(package_match.group(1))
        params['package_width'] = int(package_match.group(2))
        params['package_height'] = int(package_match.group(3))

    # Вес
    weight_match = re.search(r'Вес брутто:\s*(\d+)\s*г', desc)
    if weight_match:
        weight_g = int(weight_match.group(1))
        params['package_weight'] = round(weight_g / 1000, 3)
        params['product_weight'] = round(weight_g / 1000, 3)

    # Покрытие
    coating_match = re.search(r'Покрытие:\s*([^\n]+)', desc)
    if coating_match:
        params['coating'] = coating_match.group(1).strip()

    # Страна
    country_match = re.search(r'Страна производитель:\s*([^\n]+)', desc)
    if country_match:
        params['country'] = country_match.group(1).strip()

    # Присоединительный размер
    connection_match = re.search(r'Присоединительный размер:\s*([^\n]+)', desc)
    if connection_match:
        params['connection_size'] = connection_match.group(1).strip()

    # Тип излива
    type_match = re.search(r'Тип излива:\s*([^\n]+)', desc)
    if type_match:
        params['type'] = type_match.group(1).strip()

    # Монтаж
    mount_match = re.search(r'Монтаж:\s*([^\n]+)', desc)
    if mount_match:
        params['mount_type'] = mount_match.group(1).strip()

    # Подводка
    hose_match = re.search(r'Подводка:\s*([^\n]+)', desc)
    if hose_match:
        params['hose_type'] = hose_match.group(1).strip()

    return params

def create_accurate_description(name, model, params, original_desc):
    """Создает точное описание товара на основе его характеристик"""

    name_lower = str(name).lower()
    model_str = str(model) if pd.notna(model) else ""

    # СМЕСИТЕЛИ ДЛЯ РАКОВИНЫ
    if 'смеситель' in name_lower and 'раковин' in name_lower:
        desc = f"Смеситель для раковины"

        # Добавляем тип по модели
        if 'шпилька' in model_str.lower():
            desc += " с креплением на шпильку"
        elif 'гайка' in model_str.lower():
            desc += " с креплением на гайку"

        desc += ". "

        # Тип излива
        if params['type']:
            if 'фиксированный' in params['type'].lower():
                desc += "Фиксированный излив не поворачивается, обеспечивает стабильную подачу воды. "
            elif 'поворотный' in params['type'].lower():
                desc += "Поворотный излив увеличивает радиус действия и удобство использования. "

        # Покрытие
        if params['coating']:
            if 'хром' in params['coating'].lower():
                desc += "Хромированное покрытие придает блеск и защищает от коррозии. "

        # Размеры
        if params['product_length']:
            desc += f"Длина излива {params['product_length']} мм. "

        # Подключение
        if params['connection_size']:
            desc += f"Стандартное подключение {params['connection_size']}. "

        if params['hose_type']:
            desc += f"Подводка: {params['hose_type'].lower()}. "

    # СМЕСИТЕЛИ ДЛЯ МОЙКИ
    elif 'смеситель' in name_lower and 'мойк' in name_lower:
        desc = f"Смеситель для кухонной мойки"

        if 'гайка' in model_str.lower():
            desc += " с креплением на гайку"
        elif 'шпилька' in model_str.lower():
            desc += " с креплением на шпильку"

        desc += ". "

        # Тип излива
        if params['type']:
            if 'поворотный' in params['type'].lower():
                desc += "Поворотный излив обеспечивает максимальное удобство при мытье посуды. "
            elif 'фиксированный' in params['type'].lower():
                desc += "Фиксированный излив для стабильной подачи воды. "

        # Размеры
        if params['product_length']:
            desc += f"Длина излива {params['product_length']} мм для удобного доступа к раковине. "

        # Покрытие
        if params['coating'] and 'хром' in params['coating'].lower():
            desc += "Хромированная поверхность легко очищается и не тускнеет. "

    # СМЕСИТЕЛИ ДЛЯ УМЫВАЛЬНИКА
    elif 'смеситель' in name_lower and 'умывальник' in name_lower:
        desc = f"Смеситель для умывальника"

        if 'шпилька' in model_str.lower():
            desc += " с креплением на шпильку"
        elif 'гайка' in model_str.lower():
            desc += " с креплением на гайку"

        desc += ". "

        if params['type']:
            if 'поворотный' in params['type'].lower():
                desc += "Поворотный механизм излива для максимального комфорта. "

        if params['product_length']:
            desc += f"Длина излива {params['product_length']} мм. "

    # РАСПЫЛИТЕЛИ
    elif 'распылитель' in name_lower:
        if 'веерный' in name_lower:
            desc = "Веерный распылитель создает широкий поток воды в форме веера для равномерного полива растений. "
        elif 'импульсный' in name_lower:
            desc = "Импульсный распылитель обеспечивает дальнобойный полив за счет пульсирующих струй воды. "
        else:
            desc = "Распылитель для садового полива. "

        # Цвет по модели
        if 'зеленый' in model_str.lower():
            desc += "Зеленый цвет корпуса. "
        elif 'синий' in model_str.lower():
            desc += "Синий цвет корпуса. "

    # ДОЖДЕВАТЕЛИ
    elif 'дождеватель' in name_lower:
        if 'круговой' in name_lower:
            desc = "Круговой дождеватель обеспечивает равномерное орошение по окружности. "
        elif 'осциллирующий' in name_lower:
            desc = "Осциллирующий дождеватель покрывает прямоугольную площадь маятниковыми движениями. "
        else:
            desc = "Дождеватель для автоматического полива участка. "

    # НАСАДКИ И ПИСТОЛЕТЫ
    elif 'насадка' in name_lower or 'пистолет' in name_lower:
        desc = "Поливочная насадка с регулировкой режимов распыления. "
        if 'многофункциональная' in name_lower:
            desc += "Несколько режимов полива от тонкой струи до мягкого распыления. "

    else:
        # Универсальное описание
        desc = f"Качественное изделие для водоснабжения и полива. "

    # Общие характеристики
    if params['country']:
        desc += f"Производство: {params['country']}. "

    return desc.strip()

def create_additional_description(name, params):
    """Создает дополнительное описание"""
    details = []

    if params['coating']:
        details.append(f"Покрытие: {params['coating']}")

    if params['mount_type']:
        details.append(f"Тип монтажа: {params['mount_type'].lower()}")

    if params['connection_size']:
        details.append(f"Присоединение: {params['connection_size']}")

    if params['product_length'] and params['product_height']:
        details.append(f"Габариты: {params['product_length']}×{params['product_height']} мм")
    elif params['product_length']:
        details.append(f"Длина: {params['product_length']} мм")

    if params['package_length'] and params['package_width'] and params['package_height']:
        details.append(f"Упаковка: {params['package_length']}×{params['package_width']}×{params['package_height']} мм")

    if len(details) == 0:
        return "Соответствует стандартам качества."

    return ". ".join(details) + "."

# Обрабатываем все товары
processed_products = []

print("=== СОЗДАНИЕ ИНДИВИДУАЛЬНЫХ ОПИСАНИЙ ===")

for idx in range(3, len(products_df)):
    product_row = products_df.iloc[idx]

    if pd.isna(product_row[2]):
        continue

    params = parse_product_description(product_row[4])

    name = str(product_row[2]) if pd.notna(product_row[2]) else ''
    model = str(product_row[3]) if pd.notna(product_row[3]) else ''

    description = create_accurate_description(name, model, params, product_row[4])
    additional_desc = create_additional_description(name, params)

    product = {
        'name': name,
        'model': model,
        'article': str(product_row[5]) if pd.notna(product_row[5]) else '',
        'original_description': str(product_row[4]) if pd.notna(product_row[4]) else '',
        'parameters': params,
        'description': description,
        'additional_description': additional_desc
    }

    processed_products.append(product)

    # Показываем прогресс
    print(f"{len(processed_products):2d}. {name[:50]}")
    print(f"    Описание: {description[:80]}...")
    print()

print(f"\nОбработано {len(processed_products)} товаров")

# Создаем финальный Excel
wb = Workbook()
ws = wb.active
ws.title = "Товары для Петрович"

headers = [
    "Наименование товара от поставщика",
    "Описание товара",
    "Дополнительное описание (Необязательное)",
    "Страна происхождения",
    "Код альтозиции «Петрович»",
    "Преимущество перед аналогом",
    "Ссылка на товар на вашем сайте",
    "Базовая единица измерения",
    "Длина, мм.",
    "Ширина, мм.",
    "Высота, мм",
    "Вес, кг",
    "Длина изделия, мм.",
    "Ширина изделия, мм.",
    "Высота изделия, мм",
    "Вес изделия, кг",
    "Контроль цены для покупателя (да/нет)",
    "Вид упаковки",
    "Штрих код",
    "Тип штрих кода"
]

# Заголовки
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Данные
for row_idx, product in enumerate(processed_products, 2):
    params = product['parameters']

    ws.cell(row=row_idx, column=1, value=product['name'])
    ws.cell(row=row_idx, column=2, value=product['description'])
    ws.cell(row=row_idx, column=3, value=product['additional_description'])
    ws.cell(row=row_idx, column=4, value=params.get('country', 'Россия'))
    ws.cell(row=row_idx, column=5, value=product['article'])
    ws.cell(row=row_idx, column=6, value="Надежное качество, оптимальная цена, простота установки")
    ws.cell(row=row_idx, column=7, value="")
    ws.cell(row=row_idx, column=8, value="шт")

    # Размеры упаковки
    ws.cell(row=row_idx, column=9, value=params['package_length'])
    ws.cell(row=row_idx, column=10, value=params['package_width'])
    ws.cell(row=row_idx, column=11, value=params['package_height'])
    ws.cell(row=row_idx, column=12, value=params['package_weight'])

    # Размеры изделия
    ws.cell(row=row_idx, column=13, value=params['product_length'])
    ws.cell(row=row_idx, column=14, value=params['product_width'])
    ws.cell(row=row_idx, column=15, value=params['product_height'])
    ws.cell(row=row_idx, column=16, value=params['product_weight'])

    ws.cell(row=row_idx, column=17, value="да")
    ws.cell(row=row_idx, column=18, value="Картонная коробка")
    ws.cell(row=row_idx, column=19, value="")
    ws.cell(row=row_idx, column=20, value="EAN13")

# Автоподбор ширины
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 60)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save('template_petrovich_FINAL.xlsx')

# Сохраняем JSON
with open('final_products.json', 'w', encoding='utf-8') as f:
    json.dump(processed_products, f, ensure_ascii=False, indent=2)

print("\nГОТОВО!")
print("template_petrovich_FINAL.xlsx - финальный Excel с правильными описаниями")
print("final_products.json - JSON с обработанными данными")